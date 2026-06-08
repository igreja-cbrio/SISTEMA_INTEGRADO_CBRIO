"""
Gera docs/permissoes-mapa.xlsx · planilha pro Marcos preencher a matriz
de permissões por cargo + overrides por pessoa.

Marcos pediu xlsx (em vez de markdown) pra facilitar o preenchimento.

Como rodar:
    python scripts/gerar_permissoes_xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Estilos comuns
# ---------------------------------------------------------------------------
CORES = {
    "primario":   "00B39D",
    "primario_bg":"E0F5F2",
    "header_bg":  "1F2937",
    "header_fg":  "FFFFFF",
    "sub_bg":     "F3F4F6",
    "estrategico":"DBEAFE",
    "ministerial":"FAE8FF",
    "operacoes":  "FEF3C7",
    "dados":      "DCFCE7",
    "admin":      "FECACA",
    "perg_bg":    "FFF7ED",
}


def borda_fina():
    side = Side(style="thin", color="D1D5DB")
    return Border(left=side, right=side, top=side, bottom=side)


def borda_grossa():
    side = Side(style="medium", color="111827")
    return Border(left=side, right=side, top=side, bottom=side)


def aplicar_header(cel, fill_hex=CORES["header_bg"], fg_hex=CORES["header_fg"], wrap=True):
    cel.font = Font(name="Calibri", size=11, bold=True, color=fg_hex)
    cel.fill = PatternFill("solid", fgColor=fill_hex)
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cel.border = borda_fina()


def aplicar_celula(cel, wrap=True, h="left", v="top"):
    cel.font = Font(name="Calibri", size=10)
    cel.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    cel.border = borda_fina()


def aplicar_fill(cel, hex_color):
    cel.fill = PatternFill("solid", fgColor=hex_color)


def setar_largura(ws, larguras):
    for col_idx, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ---------------------------------------------------------------------------
# Dados (fonte única · usado nas várias abas)
# ---------------------------------------------------------------------------

CARGOS = [
    # (nome curto pra header da matriz, nome completo, titular sugerido)
    ("Pastor Sr",     "Pastor Senior",                       "Pr. Pedrão"),
    ("Pastor Pres",   "Pastor Presidente",                   "Pr. Juninho"),
    ("Dir Geral",     "Diretor Geral / CEO",                 "Eduardo Gnisci"),
    ("Dir Estrat",    "Diretor de Estratégia (PMO)",         "Marcos Paulo"),
    ("Líder Mini",    "Líder Ministerial",                   "Arthur Serpa"),
    ("Líder Criat",   "Líder Criativo",                      "Pedro Menezes"),
    ("Líder Área",    "Líder de Área Ministerial",           "Alda (Integração) + líderes AMI/Bridge/Sede/Online/Kids/Cuidados/Voluntariado/Next/Grupos/Generosidade/CBA"),
    ("Assist Área",   "Assistente de Área",                  "Mão direita do líder"),
    ("Líder Fin",     "Líder Financeiro",                    "Yago Torres"),
    ("Líder Mkt",     "Líder de Marketing",                  "Pedro Paiva"),
    ("Líder Prod",    "Líder de Produção",                   "Pedro Fernandes"),
    ("Líder Op",      "Líder de Operações (Hospitalidade)",  "Jéssica Salviano · Amaury (cozinha/limpeza/manutenção/compras)"),
    ("Líder RH",      "Líder de RH",                         "vago"),
    ("Coord Vol",     "Coordenador de Voluntários",          "?"),
    ("Voluntário",    "Voluntário",                          "qualquer pessoa que serve"),
    ("Membro",        "Membro",                              "auto-cadastro · dashboard básico"),
    ("Conselho",      "Conselho Estatutário",                "não-funcionário · vê dashboards"),
    ("Dev",           "Suporte/Dev",                         "Matheus + Marcos"),
]

# Lista de módulos agrupada por matriz (chave = nome da aba)
MODULOS_POR_MATRIZ = {
    "Matriz Estratégica": [
        ("Dashboard",              "/dashboard",            "Home com cards resumo"),
        ("Painel CBRio",           "/painel",               "NSM · mandalas · matrizes · alertas"),
        ("Minha Área",             "/minha-area",           "KPIs do líder (filtrado por área/valor)"),
        ("Gestão (PMO)",           "/gestao",               "Estrutura OKR · configurar metas · saúde sistema"),
        ("Planejamento",           "/planejamento",         "Ritual mensal causa-decisão"),
        ("Ritual Mensal",          "/ritual",               "Revisão da Diretoria Geral (5 nominais)"),
        ("Governança",             "/governanca",           "Ciclo mensal OKR · DRE · KPI · Conselho"),
        ("Revisão Estratégica",    "/revisao-estrategica",  "Edição direta de projetos/marcos · cascata impacto"),
    ],
    "Matriz Ministerial": [
        ("Integração",             "/ministerial/integracao",      "Cultos · Frequência · Decisões · Batismos · Histórico"),
        ("Cuidados",               "/ministerial/cuidados",        "Acompanhamentos pastorais · Jornada 180 · Convertidos"),
        ("Online (YouTube)",       "/ministerial/online",          "Desempenho do canal (read-only)"),
        ("NEXT",                   "/ministerial/next",            "Curso de novos membros"),
        ("Voluntariado",           "/ministerial/voluntariado/*",  "Checkin · escalas · perfil · disponibilidade"),
        ("Membresia",              "/ministerial/membresia",       "CRM de pessoas · jornada · cartão digital"),
        ("Grupos",                 "/grupos",                      "Grupos de conexão · supervisão · pedidos"),
    ],
    "Matriz Operações": [
        ("Eventos",                "/eventos",          "Ciclo criativo · fases · documentos · KPIs por evento"),
        ("Projetos",               "/projetos",         "Projetos com fases"),
        ("Expansão",               "/expansao",         "Marcos estratégicos até 2029"),
        ("Processos",              "/processos",        "Processos operacionais que alimentam KPIs"),
        ("RH",                     "/admin/rh",         "Funcionários · documentos · treinamentos"),
        ("Financeiro",             "/admin/financeiro", "Receitas · despesas · relatórios"),
        ("Logística",              "/admin/logistica",  "Estoque · compras · almoxarifado"),
        ("Patrimônio",             "/admin/patrimonio", "Espaços · equipamentos · inventário"),
        ("Solicitações",           "/solicitacoes",     "Backbone administrativo · SLA · aprovações"),
    ],
    "Matriz Dados / IA / Admin": [
        ("Dados Brutos",           "/dados-brutos",       "Líder preenche números absolutos"),
        ("NPS",                    "/nps",                "Pesquisas · respostas · link público"),
        ("Notificações (config)",  "(admin config)",      "Regras de quem recebe alertas de cada módulo"),
        ("Assistente IA",          "/assistente-ia",      "Agente Claude conversacional"),
        ("Cérebro CBRio",          "(backend cron)",      "Sync SharePoint → Obsidian via Haiku"),
        ("Perfil próprio",         "/perfil",             "Dados pessoais do próprio usuário"),
        ("Permissões (admin)",     "/admin/permissoes",   "UI deste sistema · gestão de cargos + overrides"),
        ("Usuários (admin)",       "/admin/usuarios",     "Cadastrar/desativar pessoas"),
    ],
}

CORES_DAS_MATRIZES = {
    "Matriz Estratégica":         CORES["estrategico"],
    "Matriz Ministerial":         CORES["ministerial"],
    "Matriz Operações":           CORES["operacoes"],
    "Matriz Dados / IA / Admin":  CORES["dados"],
}


# ---------------------------------------------------------------------------
# Abas
# ---------------------------------------------------------------------------

def aba_capa(wb):
    ws = wb.active
    ws.title = "📋 Comece aqui"
    setar_largura(ws, [4, 110, 30])

    ws["B2"] = "Mapa de Permissões CBRio · planilha pra preencher e devolver"
    ws["B2"].font = Font(name="Calibri", size=20, bold=True, color="00B39D")
    ws.merge_cells("B2:C2")

    instrucoes = [
        "",
        "Como funciona esta planilha:",
        "",
        "1. Leia o Modelo proposto abaixo · confirme se faz sentido.",
        "2. Revise os Cargos (aba 🪪 Cargos) · adicione/remova/renomeie.",
        "3. Revise os Módulos (aba 📚 Módulos) · confirme que cobre tudo.",
        "4. Preencha as 4 matrizes (abas ⭐, ⛪, 🔧, 📊) com a legenda 0-5.",
        "5. Responda as 10 perguntas em aberto (aba ❓ Perguntas).",
        "6. Me devolva a planilha · eu implemento o que ficou definido.",
        "",
        "Pode escrever '?' onde tiver dúvida · eu volto e proponho um valor.",
        "",
        "—————————————————————————————————————————————————————",
        "",
        "Modelo proposto (você definiu):",
        "",
        "Permissão padrão = por cargo. Pessoa que ocupa cargo herda o pacote.",
        "Pessoa sai do cargo, perde o pacote. Próxima pessoa herda igual.",
        "",
        "Override por pessoa = exceção. Quando alguém precisa acessar algo",
        "fora do seu pacote (cobrir licença, projeto pontual, etc.),",
        "administrador concede individualmente · ideal com data de expiração.",
        "",
        "Hierarquia em camadas (permissão efetiva = soma):",
        "    cargo + overrides 'adicionar' − overrides 'remover'",
        "",
        "—————————————————————————————————————————————————————",
        "",
        "Legenda dos níveis (use nos campos da matriz):",
        "",
        "0 — Sem acesso · módulo não aparece no menu nem responde a URL",
        "1 — Ver · só leitura, sem editar nem exportar",
        "2 — Ver + preencher dado bruto · pode lançar números na sua área",
        "3 — Ver + editar · CRUD do conteúdo (criar, alterar)",
        "4 — Ver + editar + deletar",
        "5 — Admin do módulo · configura regras, metas, seeds, deleta tudo",
        "",
        "Modificadores opcionais (escreva depois do número, ex: 3+E ou 4*):",
        "+E — pode exportar dados (LGPD · CPF, telefone, financeiro)",
        "+A — pode aprovar workflows daquele módulo (ex: aprovar despesa)",
        "*  — acesso só da própria área (ex: líder AMI só edita cultos AMI)",
        "?  — indeciso · me pede sugestão",
        "",
        "Exemplos:",
        "    Líder Integração na Integração:  3*    (CRUD + só da área)",
        "    Diretor Geral em Financeiro:     5+E+A",
        "    Voluntário em Voluntariado:      2",
    ]
    for i, txt in enumerate(instrucoes, start=4):
        c = ws.cell(row=i, column=2, value=txt)
        if txt.startswith("—"):
            c.font = Font(name="Calibri", size=10, color="9CA3AF")
        elif txt.endswith(":") and txt.strip():
            c.font = Font(name="Calibri", size=12, bold=True, color="111827")
        elif txt.strip().startswith("Modelo proposto"):
            c.font = Font(name="Calibri", size=12, bold=True, color="111827")
        elif len(txt) > 0 and (txt[0].isdigit() and txt[1] in (" ", "—", ".")):
            c.font = Font(name="Calibri", size=10)
        else:
            c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")


def aba_cargos(wb):
    ws = wb.create_sheet("🪪 Cargos")
    setar_largura(ws, [4, 5, 32, 60, 18, 30])

    ws["B2"] = "Catálogo de cargos sugerido"
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color="00B39D")
    ws.merge_cells("B2:F2")

    ws["B3"] = (
        "Baseado no que vi no código + CLAUDE.md + memória. "
        "Confirma cada um e adiciona o que faltar. "
        "Pra cada cargo, opcionalmente diz o titular atual (facilita o seed depois)."
    )
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B3"].font = Font(name="Calibri", size=10, color="6B7280")
    ws.merge_cells("B3:F3")
    ws.row_dimensions[3].height = 35

    headers = ["#", "Cargo", "Titular atual sugerido", "Confirma? (✔ / ✘)", "Renomear? (escreva novo nome)"]
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=5, column=col, value=h)
        aplicar_header(c)
    ws.row_dimensions[5].height = 30

    for i, (curto, longo, titular) in enumerate(CARGOS):
        row = 6 + i
        ws.cell(row=row, column=2, value=i + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=longo)
        ws.cell(row=row, column=4, value=titular)
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value="")
        for col in range(2, 7):
            cel = ws.cell(row=row, column=col)
            aplicar_celula(cel)
        ws.row_dimensions[row].height = 30

    # 2 linhas em branco pra adicionar cargos novos
    n = 6 + len(CARGOS)
    for j in range(2):
        for col in range(2, 7):
            cel = ws.cell(row=n + j, column=col, value="" if col != 2 else len(CARGOS) + 1 + j)
            aplicar_celula(cel)
            aplicar_fill(cel, "FFF7ED")
        ws.row_dimensions[n + j].height = 30


def aba_modulos(wb):
    ws = wb.create_sheet("📚 Módulos")
    setar_largura(ws, [4, 28, 30, 65, 25])

    ws["B2"] = "Catálogo de módulos do sistema"
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color="00B39D")
    ws.merge_cells("B2:E2")

    ws["B3"] = (
        "Lista do que existe hoje em produção. "
        "Confirma que cobre tudo. Se faltar algum, adiciona nas linhas em laranja no fim."
    )
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B3"].font = Font(name="Calibri", size=10, color="6B7280")
    ws.merge_cells("B3:E3")
    ws.row_dimensions[3].height = 30

    headers = ["Módulo", "Rota", "O que faz", "Confirma? (✔ / ✘)"]
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=5, column=col, value=h)
        aplicar_header(c)
    ws.row_dimensions[5].height = 30

    row = 6
    for matriz_nome, modulos in MODULOS_POR_MATRIZ.items():
        # Header da matriz
        cel = ws.cell(row=row, column=2, value=matriz_nome.replace("Matriz ", "").upper())
        cel.font = Font(name="Calibri", size=11, bold=True, color="111827")
        aplicar_fill(cel, CORES_DAS_MATRIZES[matriz_nome])
        cel.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 22
        row += 1

        for nome, rota, desc in modulos:
            ws.cell(row=row, column=2, value=nome).font = Font(name="Calibri", size=10, bold=True)
            ws.cell(row=row, column=3, value=rota).font = Font(name="Consolas", size=9, color="6B7280")
            ws.cell(row=row, column=4, value=desc)
            ws.cell(row=row, column=5, value="")
            for col in range(2, 6):
                cel = ws.cell(row=row, column=col)
                if col == 2:
                    cel.alignment = Alignment(vertical="center", wrap_text=True)
                else:
                    aplicar_celula(cel)
            ws.row_dimensions[row].height = 25
            row += 1

    # Linhas em laranja pra adicionar módulos novos
    for j in range(3):
        for col in range(2, 6):
            cel = ws.cell(row=row + j, column=col, value="")
            aplicar_celula(cel)
            aplicar_fill(cel, "FFF7ED")
        ws.row_dimensions[row + j].height = 25


def aba_matriz(wb, nome, modulos, cor_bg):
    ws = wb.create_sheet(nome)

    # 1 col índice + N cols pra cargos
    n_cargos = len(CARGOS)
    larguras = [4, 26] + [12] * n_cargos
    setar_largura(ws, larguras)

    # Título
    ws["B2"] = nome
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color="00B39D")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=2 + n_cargos)

    # Subtítulo
    ws["B3"] = (
        "Preencha cada célula com a legenda 0-5 (veja '📋 Comece aqui'). "
        "Modificadores opcionais: +E (exportar), +A (aprovar), * (só própria área), ? (sugere)."
    )
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B3"].font = Font(name="Calibri", size=9, color="6B7280")
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=2 + n_cargos)
    ws.row_dimensions[3].height = 28

    # Header: módulo + nomes curtos dos cargos
    row_h = 5
    cel = ws.cell(row=row_h, column=2, value="Módulo")
    aplicar_header(cel)
    for i, (curto, longo, titular) in enumerate(CARGOS):
        cel = ws.cell(row=row_h, column=3 + i, value=curto)
        aplicar_header(cel)
        cel.comment = None  # mantém limpo · nome longo + titular em outra aba

    ws.row_dimensions[row_h].height = 50

    # Linhas dos módulos
    for r, (nome_mod, rota, desc) in enumerate(modulos):
        row = row_h + 1 + r
        cel_mod = ws.cell(row=row, column=2, value=nome_mod)
        cel_mod.font = Font(name="Calibri", size=10, bold=True)
        cel_mod.fill = PatternFill("solid", fgColor=cor_bg)
        cel_mod.alignment = Alignment(vertical="center", wrap_text=True)
        cel_mod.border = borda_fina()

        for i in range(n_cargos):
            cel = ws.cell(row=row, column=3 + i, value="")
            cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.border = borda_fina()
            cel.font = Font(name="Calibri", size=11, bold=True)

        ws.row_dimensions[row].height = 26

    # Congelar painéis (header de cargos + coluna de módulo sempre visíveis)
    ws.freeze_panes = "C6"


def aba_overrides(wb):
    ws = wb.create_sheet("⚙️ Overrides por pessoa")
    setar_largura(ws, [4, 26, 22, 12, 8, 18, 45, 22, 14])

    ws["B2"] = "Overrides por pessoa · exceções"
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color="00B39D")
    ws.merge_cells("B2:I2")

    ws["B3"] = (
        "Quando alguém precisa acessar algo fora do seu cargo (cobrir licença, projeto pontual, etc.). "
        "Preencha exemplos reais que vc consegue prever · me ajuda a calibrar o sistema. "
        "Pode deixar em branco se preferir definir depois."
    )
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B3"].font = Font(name="Calibri", size=10, color="6B7280")
    ws.merge_cells("B3:I3")
    ws.row_dimensions[3].height = 40

    headers = ["Pessoa (nome)", "Módulo", "Tipo (adicionar/remover)", "Nível 0-5", "Modificadores (+E/+A/*)", "Motivo (texto livre)", "Concedido por (nome)", "Válido até (data ou 'indefinido')"]
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=5, column=col, value=h)
        aplicar_header(c)
    ws.row_dimensions[5].height = 40

    # Exemplos pré-preenchidos
    exemplos = [
        ("Marcos Paulo",      "Financeiro",   "adicionar", "5", "+E +A", "Cobertura enquanto Líder Financeiro está em licença", "Eduardo Gnisci", "2026-07-15"),
        ("Voluntário X",      "Cuidados",     "adicionar", "2", "",      "Líder informal de grupo · ajuda no follow-up",          "Alda Lorena",    "indefinido"),
        ("Líder Y",           "Logística",    "remover",   "0", "",      "Saiu por conflito · perde acesso mas mantém cargo",     "Eduardo Gnisci", "indefinido"),
        ("Estagiário Z",      "Membresia",    "adicionar", "1", "",      "Pode ver lista durante onboarding",                      "RH",             "2026-08-01"),
    ]
    for i, ex in enumerate(exemplos):
        row = 6 + i
        for col, v in enumerate(ex, start=2):
            cel = ws.cell(row=row, column=col, value=v)
            aplicar_celula(cel)
            aplicar_fill(cel, "F3F4F6")
        ws.row_dimensions[row].height = 32

    # Linhas em branco pro Marcos preencher
    for j in range(15):
        row = 6 + len(exemplos) + j
        for col in range(2, 10):
            cel = ws.cell(row=row, column=col, value="")
            aplicar_celula(cel)
        ws.row_dimensions[row].height = 28

    # Decisões sobre overrides
    ws.cell(row=6 + len(exemplos) + 16, column=2, value="Decisões sobre overrides:").font = Font(name="Calibri", size=12, bold=True, color="111827")
    decisoes = [
        ("Override expira automaticamente ou fica indefinido por padrão?", "[ ] 30 dias  [ ] 90 dias (sugiro)  [ ] indefinido"),
        ("Quem pode conceder override?",                                    "[ ] só admin/diretor  [ ] admin/diretor + líder na sua área (sugiro)  [ ] qualquer cargo no seu escopo"),
        ("Queremos histórico/auditoria das concessões?",                    "[ ] sim, tabela permissao_log (sugiro)  [ ] não"),
    ]
    for i, (perg, opts) in enumerate(decisoes):
        r = 6 + len(exemplos) + 18 + i * 2
        ws.cell(row=r, column=2, value=perg).font = Font(name="Calibri", size=10, bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
        cel = ws.cell(row=r + 1, column=2, value=opts)
        cel.font = Font(name="Calibri", size=10)
        ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=9)
        ws.row_dimensions[r].height = 22
        ws.row_dimensions[r + 1].height = 22


def aba_perguntas(wb):
    ws = wb.create_sheet("❓ Perguntas em aberto")
    setar_largura(ws, [4, 5, 65, 50])

    ws["B2"] = "Perguntas em aberto · marca uma opção em cada"
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color="00B39D")
    ws.merge_cells("B2:D2")

    ws["B3"] = "Responder rápido com [X] na opção que prefere. Onde tiver dúvida, deixa '?' que eu volto e proponho."
    ws["B3"].font = Font(name="Calibri", size=10, color="6B7280")
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("B3:D3")
    ws.row_dimensions[3].height = 25

    headers = ["#", "Pergunta", "Opções (marque uma com X)"]
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=5, column=col, value=h)
        aplicar_header(c)
    ws.row_dimensions[5].height = 28

    perguntas = [
        ("Pessoas com múltiplos cargos (ex: Marcos = Diretor Estratégia + área 'adm'):",
         "[ ] múltiplos cargos somando perms\n[ ] cargo único combinado"),
        ("Cargo 'Líder de Área' genérico ou específico por área?",
         "[ ] genérico vinculado à área (mais simples)\n[ ] específico por área (mais explícito)"),
        ("Pastor Sênior · 100% leitura ou pode tudo?",
         "[ ] leitura total + ritual (observador)\n[ ] admin tudo"),
        ("Conselho Estatutário · só dashboard executivo?",
         "[ ] sim (1 em Painel/Dashboard/Gestão, 0 no resto)\n[ ] outro arranjo: _______"),
        ("Exportar dados sensíveis (+E) · quem pode tirar relatório com CPF/telefone?",
         "[ ] Dir Geral + Dir Estrat + Líderes da área do dado\n[ ] outro: _______"),
        ("Voluntário ganha acesso só após onboarding (admin promove de '0 em tudo' → 'Voluntário ativo')?",
         "[ ] sim, fluxo de onboarding\n[ ] já entra como Voluntário direto"),
        ("Dev (Matheus/Marcos como dev) · cargo único 'Dev' com 5+E+A em tudo?",
         "[ ] full em tudo\n[ ] separar Dev backend / Dev frontend"),
        ("Override por pessoa · quem concede?",
         "[ ] só admin/diretor\n[ ] admin/diretor + líder na sua área (sugiro)\n[ ] livre dentro do escopo"),
        ("Expiração padrão do override:",
         "[ ] 30 dias\n[ ] 90 dias (sugiro)\n[ ] indefinido"),
        ("Migração: mapear `profiles.role` atual pra cargo equivalente · manter `kpi_areas`/`kpi_valores` deprecated por 1 release · substituir middlewares ad-hoc por `authorize(modulo, acao)`. OK?",
         "[ ] OK\n[ ] quero outro caminho: _______"),
    ]
    for i, (perg, opts) in enumerate(perguntas):
        row = 6 + i
        ws.cell(row=row, column=2, value=i + 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row, column=2).border = borda_fina()

        cel_p = ws.cell(row=row, column=3, value=perg)
        cel_p.alignment = Alignment(wrap_text=True, vertical="top")
        cel_p.font = Font(name="Calibri", size=10)
        cel_p.border = borda_fina()

        cel_o = ws.cell(row=row, column=4, value=opts)
        cel_o.alignment = Alignment(wrap_text=True, vertical="top")
        cel_o.font = Font(name="Calibri", size=10)
        cel_o.fill = PatternFill("solid", fgColor=CORES["perg_bg"])
        cel_o.border = borda_fina()

        ws.row_dimensions[row].height = 52


def aba_referencia(wb):
    ws = wb.create_sheet("📖 Referência de cargos")
    setar_largura(ws, [4, 16, 36, 60])

    ws["B2"] = "Referência rápida dos cargos (com nome longo + titular)"
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color="00B39D")
    ws.merge_cells("B2:D2")

    ws["B3"] = "Quando preencher as matrizes, use o nome curto da coluna · aqui está o nome completo + titular."
    ws["B3"].font = Font(name="Calibri", size=10, color="6B7280")
    ws.merge_cells("B3:D3")
    ws.row_dimensions[3].height = 22

    headers = ["Cód", "Nome curto (matriz)", "Nome completo", "Titular sugerido"]
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=5, column=col, value=h)
        aplicar_header(c)
    ws.row_dimensions[5].height = 28

    for i, (curto, longo, titular) in enumerate(CARGOS):
        row = 6 + i
        ws.cell(row=row, column=2, value=i + 1)
        ws.cell(row=row, column=3, value=curto).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=row, column=4, value=longo)
        ws.cell(row=row, column=5, value=titular)
        for col in range(2, 6):
            cel = ws.cell(row=row, column=col)
            aplicar_celula(cel)
        ws.row_dimensions[row].height = 25


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    aba_capa(wb)
    aba_referencia(wb)
    aba_cargos(wb)
    aba_modulos(wb)

    icones_matriz = ["⭐ Matriz Estratégica", "⛪ Matriz Ministerial", "🔧 Matriz Operações", "📊 Matriz Dados-IA-Admin"]
    chaves_matriz = list(MODULOS_POR_MATRIZ.keys())
    for icone, chave in zip(icones_matriz, chaves_matriz):
        aba_matriz(wb, icone, MODULOS_POR_MATRIZ[chave], CORES_DAS_MATRIZES[chave])

    aba_overrides(wb)
    aba_perguntas(wb)

    out = "docs/permissoes-mapa.xlsx"
    wb.save(out)
    print(f"OK · {out}")


if __name__ == "__main__":
    main()
